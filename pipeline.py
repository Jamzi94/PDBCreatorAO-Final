from __future__ import annotations

import argparse
import json
import math
import re
import sys
import threading
import warnings

import numpy as np

# Suppress widespread BiopythonDeprecationWarning from command-line wrappers
from Bio import BiopythonDeprecationWarning
warnings.filterwarnings("ignore", category=BiopythonDeprecationWarning)
import logging
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from alignment import align_and_patch
from alphafold import (
    download_alphafold_pdb,
    get_alphafold_details,
    get_alphafold_pdb_info,
)
from config import PipelineConfig
from networking import configure_retry, set_http2_mode
from oriented import align_structures_by_groups
from pdb_details import get_pdb_details
from pdb_download import download_pdb_structure
from structure_analysis import (
    detect_gaps_and_discard_coils,
    parse_helix_sheet_annotations,
    parse_small_molecules,
)
from structure_repair import fix_structure
from uniprot import fetch_uniprot_json, find_related_uniprots, parse_uniprot_data

_WORKER_LOG_FILE: Optional[Path] = None

def _configure_worker_logging() -> None:
    """Configure logging for worker processes to write to both console and file."""
    global _WORKER_LOG_FILE
    root = logging.getLogger()
    if root.handlers:
        return

    # If the global log file path is set, use it to configure file and stream handlers.
    if _WORKER_LOG_FILE:
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[
                logging.FileHandler(_WORKER_LOG_FILE, "a", "utf-8"),  # Use append mode
                logging.StreamHandler(),
            ],
            force=True,
        )
    else:  # Fallback to console-only if the path isn't set
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
        )


def _alignment_worker(task: tuple[str, str, str, int, int, str, list[str], str]) -> dict:
    _configure_worker_logging()
    af_path, exp_path, gene_name, prot_len, window, outdir, partner_ids, uniprot_id = task
    return align_and_patch(
        af_path=af_path,
        exp_path=exp_path,
        gene=gene_name,
        prot_len=prot_len,
        window=window,
        outdir=outdir,
        partner_ids=partner_ids,
        uniprot_id=uniprot_id,
    )


# --- Data loading and preparation helpers ---

HYP_RE = re.compile(
    r'=?HYPERLINK\(\s*"[^"]*"\s*[;,]\s*"(?P<id>[A-Za-z0-9\-]+)"\s*\)', re.IGNORECASE
)

def extract_pdb_from_cell(cell) -> str:
    """Extract a PDB identifier from an Excel cell."""
    if cell.hyperlink and cell.hyperlink.target:
        return Path(cell.hyperlink.target).stem.upper()
    val = cell.value
    if isinstance(val, str):
        m = HYP_RE.match(val.strip())
        return (m.group("id") if m else val).upper()
    return None if val is None else str(val).strip().upper()


def load_sheet(excel_file: Path | str, name: str) -> pd.DataFrame:
    """Load an Excel worksheet and normalise hyperlink values."""
    excel_path = Path(excel_file)
    if not excel_path.is_file():
        logging.error(f"Fichier Excel manquant: {excel_path}")
        sys.exit(1)
    wb = load_workbook(excel_path, data_only=False)
    if name not in wb.sheetnames:
        logging.error(f"Feuille '{name}' manquante ({wb.sheetnames})")
        sys.exit(1)
    ws = wb[name]
    headers = [str(c.value).strip().lower().replace(" ", "_") for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2):
        rec = {h: extract_pdb_from_cell(c) for h, c in zip(headers, row)}
        rows.append(rec)
    df = pd.DataFrame(rows)
    logging.info(f"[{name}] chargée: {df.shape[0]} lignes")
    return df


def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(
        columns=lambda c: re.sub(r"[^0-9a-zA-Z]+", "_", str(c).strip().lower())
    )


def _strip_hyperlink_value(value: Any) -> Any:
    if isinstance(value, str):
        match = HYP_RE.match(value.strip())
        if match:
            return match.group("id")
        return value.strip()
    return value


def _prepare_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    prepared = _normalise_columns(df.copy())
    for column in ("pdb_id", "uniprot_id"):
        if column in prepared.columns:
            prepared[column] = prepared[column].apply(_strip_hyperlink_value)
    return prepared


def filter_and_cascade(
    df_det: pd.DataFrame, df_ali: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Filter alignment rows to match available PDB detail entries."""
    df_det = df_det.copy()
    df_ali = df_ali.copy()
    det_ids = df_det["pdb_id"].dropna().astype(str).str.upper()
    valid = set(det_ids)
    ali_ids = df_ali["pdb_id"].fillna("").astype(str).str.upper()
    mask = ali_ids.isin(valid)
    df_ali_proc = df_ali.loc[mask].reset_index(drop=True)
    logging.info(
        "Traité: PDB_Details %d lignes, Alignment_Summary %d lignes",
        df_det.shape[0],
        df_ali_proc.shape[0],
    )
    return df_det, df_ali_proc

# --------------------------------------------------------------------------- #
# Logging setup
# --------------------------------------------------------------------------- #


def setup_logging(logfile: Path, verbose: bool) -> None:
    """Configure root logging handlers for the pipeline.

    Args:
        logfile (Path): Destination for the log file.
        verbose (bool): Enable verbose DEBUG output when True.
    """
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(logfile, "w", "utf-8"),
            logging.StreamHandler(),
        ],
        force=True,
    )


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #


def ensure_template(template: Path) -> None:
    """Create the UniProt template workbook if it does not exist.

    Args:
        template (Path): Path to the workbook to create when missing.
    """
    if template.exists():
        return
    template.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "UniProt IDs"
    ws.append(["UniProt ID"])
    wb.save(template)


def style_link(cell, link_font: Font) -> None:
    """Apply hyperlink styling when the cell contains a hyperlink formula.

    Args:
        cell (openpyxl.cell.cell.Cell): Cell to inspect and style.
        link_font (Font): Font applied to hyperlink cells.
    """
    value = cell.value
    if isinstance(value, str) and value.startswith("=HYPERLINK"):
        cell.font = link_font


# --------------------------------------------------------------------------- #
# Pipeline implementation
# --------------------------------------------------------------------------- #


class ProteinPipeline:
    """High-level orchestration of the protein processing pipeline."""

    UNIPROT_COLUMNS = [
        "UniProt ID",
        "Gene Name",
        "Protein Name",
        "PDB Entries",
        "AlphaFoldDB Entries",
        "SNPs",
        "Pharmaceutical Use",
        "TCDB ID",
        "DrugCentral ID",
        "Tissue Specificity",
        "Annotation Score",
        "Cofactors",
        "Polymorphism",
        "FASTA Sequence",
        "Orientation Methods",
    ]

    PDB_DETAILS_COLUMNS = [
        "UniProt ID",
        "Gene Name",
        "Protein Name",
        "PDB ID",
        "PDB Source",
        "Repaired",
        "Experimental Method",
        "Resolution (Angstrom)",
        "Determination Methodology",
        "Conformation",
        "Environment",
        "Polymer Entity Count",
        "Entity Descriptions",
        "Chains",
        "Entity Organisms",
        "PubMed ID",
        "Small Molecules",
    ]

    STRUCTURAL_COLUMNS = [
        "UniProt ID",
        "Gene Name",
        "Protein Name",
        "PDB ID",
        "Type",
        "Residus",
        "Length",
        "Secondary Structure",
    ]

    def __init__(
        self,
        config: PipelineConfig,
        ids: Optional[Sequence[str]] = None,
        verbose: bool = False,
    ) -> None:
        self.cfg = config
        self.ids = [uid.strip() for uid in ids] if ids else None
        self.verbose = verbose

        self.link_font = Font(color="0000FF", underline="single")
        self.log = logging.getLogger(__name__)
        self._lock = threading.Lock()

        # Accumulators
        self._uniprot_records: List[Dict[str, str]] = []
        self._pdb_detail_records: List[Dict[str, str]] = []
        self._structural_records: List[Dict[str, str]] = []
        self._alignment_summary: List[Dict[str, object]] = []
        self._multi_ref_records: List[Dict[str, object]] = []
        self._protonation_map: Dict[str, bool] = {}
        self._pdb_source_map: Dict[str, str] = {}

        self._prot_len_map: Dict[str, int] = {}
        self._partner_ids_map: Dict[str, List[str]] = {}
        self._pdb_source_map: Dict[str, str] = {}
        self._data_frames: Dict[str, pd.DataFrame] = {}

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #
    def run(self) -> None:
        """Execute the full protein processing workflow.

        Raises:
            RuntimeError: Propagated from worker tasks on fatal alignment errors.
        """
        global _WORKER_LOG_FILE
        _WORKER_LOG_FILE = self.cfg.log_file

        self.cfg.ensure_directories()
        setup_logging(self.cfg.log_file, self.verbose)
        self.log.info("Starting pipeline")
        configure_retry(
            attempts=self.cfg.api_retry.attempts,
            backoff=self.cfg.api_retry.backoff_factor,
            status=self.cfg.api_retry.status_forcelist,
        )
        set_http2_mode(self.cfg.use_http2)

        ids = self.ids or self._load_uniprot_ids()
        if not ids:
            self.log.warning("No UniProt IDs supplied; nothing to do.")
            return

        worker_count = min(len(ids), self.cfg.max_workers) or 1
        self.log.info("Using %d worker(s) for UniProt processing", worker_count)
        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            future_map = {
                executor.submit(self._process_uniprot, upid): upid for upid in ids
            }
            for future in as_completed(future_map):
                upid = future_map[future]
                try:
                    future.result()
                except Exception as exc:  # pragma: no cover - defensive logging
                    self.log.error("Processing failed for %s: %s", upid, exc)

        self._materialise_primary_frames()
        self._write_parquet_intermediates()

        self._run_alignment()
        self._materialise_alignment_frame()

        self._run_multi_reference_orientation()
        self._write_alignment_parquet()
        self._write_multi_ref_parquet()

        self._export_excel()
        self.log.info("Pipeline completed")

    # ------------------------------------------------------------------ #
    # UniProt ingestion
    # ------------------------------------------------------------------ #
    def _load_uniprot_ids(self) -> List[str]:
        ensure_template(self.cfg.input_template)
        wb = load_workbook(self.cfg.input_template)
        ws = wb.active
        ids = [
            str(row[0]).strip()
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row and row[0]
        ]
        self.log.info(
            "Loaded %d UniProt IDs from %s", len(ids), self.cfg.input_template
        )
        return ids

    def _process_uniprot(self, upid: str) -> None:
        upid = upid.strip()
        if not upid:
            return
        self.log.info("Processing %s", upid)
        try:
            up_json = fetch_uniprot_json(upid)
        except Exception as exc:
            self.log.error("UniProt fetch failed for %s: %s", upid, exc)
            return

        info = parse_uniprot_data(up_json)
        partners = find_related_uniprots(upid, up_json)

        uniprot_rows: List[Dict[str, str]] = []
        pdb_detail_rows: List[Dict[str, str]] = []
        structural_rows: List[Dict[str, str]] = []
        prot_len_updates: Dict[str, int] = {}
        partner_map: Dict[str, List[str]] = {}

        partner_ids = [p["uniprot_id"] for p in partners]
        gene = info.get("gene") or "UnknownGene"
        protein_name = info.get("protein_name", "")
        sequence = info.get("sequence", "")
        exp_ids: List[str] = info.get("pdb_ids", []) or []

        key = f"{gene}_{upid}"
        partner_map[key] = partner_ids
        prot_len_updates[key] = len(sequence)

        pred_id = get_alphafold_pdb_info(upid)
        pred_ids = [pred_id] if pred_id else []

        uniprot_rows.append(
            {
                "UniProt ID": self._hyperlink(
                    f"https://www.uniprot.org/uniprotkb/{upid}/entry", upid
                ),
                "Gene Name": self._hyperlink(
                    f"https://www.genecards.org/cgi-bin/carddisp.pl?gene={gene}", gene
                ),
                "Protein Name": protein_name,
                "PDB Entries": ", ".join(exp_ids),
                "AlphaFoldDB Entries": ", ".join(pred_ids),
                "SNPs": info.get("variants", ""),
                "Pharmaceutical Use": info.get("pharmaceutical", ""),
                "TCDB ID": info.get("tcdb_id", ""),
                "DrugCentral ID": info.get("drugcentral_id", ""),
                "Tissue Specificity": info.get("tissue", ""),
                "Annotation Score": info.get("annotation_score", ""),
                "Cofactors": info.get("cofactor", ""),
                "Polymorphism": info.get("polymorphism", ""),
                "FASTA Sequence": sequence,
                "Orientation Methods": "",
            }
        )

        exp_struct_rows, exp_pdb_rows = self._ingest_experimental_structures(
            upid=upid,
            gene=gene,
            protein_name=protein_name,
            exp_ids=exp_ids,
        )
        structural_rows.extend(exp_struct_rows)
        pdb_detail_rows.extend(exp_pdb_rows)

        af_pdb_rows, af_len_updates = self._ingest_alphafold_structures(
            upid=upid,
            gene=gene,
            protein_name=protein_name,
            pred_id=pred_id,
            partner_ids=partner_ids,
        )
        pdb_detail_rows.extend(af_pdb_rows)
        prot_len_updates.update(af_len_updates)

        with self._lock:
            self._uniprot_records.extend(uniprot_rows)
            self._pdb_detail_records.extend(pdb_detail_rows)
            self._structural_records.extend(structural_rows)
            self._prot_len_map.update(prot_len_updates)
            self._partner_ids_map.update(partner_map)

    def _ingest_experimental_structures(
        self,
        upid: str,
        gene: str,
        protein_name: str,
        exp_ids: Iterable[str],
    ) -> tuple[List[Dict[str, str]], List[Dict[str, str]]]:
        """Collect experimental structure details and return rows for downstream use."""
        structural_rows: List[Dict[str, str]] = []
        pdb_detail_rows: List[Dict[str, str]] = []

        for pdb_id in exp_ids:
            out_dir = self.cfg.exp_dir / f"{gene}_{upid}"
            out_dir.mkdir(parents=True, exist_ok=True)
            pdb_path_str, pdb_source = download_pdb_structure(
                pdb_id, str(out_dir), f"{gene}_{upid}_{pdb_id}.pdb"
            )
            if not pdb_path_str:
                continue
            pdb_source = pdb_source or ""
            pdb_path = Path(pdb_path_str)
            repaired = False
            if self.cfg.use_repair or self.cfg.use_pdb_fixer:
                try:
                    repaired = fix_structure(pdb_path, pdb_path, auto_install=self.cfg.auto_install_repair)
                    if repaired:
                        logging.info(f"PDBFixer applied to {pdb_path}")
                except Exception as e:
                    logging.error(f"PDBFixer failed for {pdb_path}: {e}")
            with self._lock:
                self._protonation_map[pdb_id.upper()] = bool(repaired)
            self._set_pdb_source(pdb_id, pdb_path, pdb_source)
            issues = detect_gaps_and_discard_coils(
                str(pdb_path),
                chain_id="A",
                plddt_threshold=60.0,
                coil_length_threshold=4,
            )
            ss_map = parse_helix_sheet_annotations(str(pdb_path), chain_id="A")
            for issue_type, start, end in issues:
                structural_rows.append(
                    {
                        "UniProt ID": self._hyperlink(
                            f"https://www.uniprot.org/uniprotkb/{upid}/entry", upid
                        ),
                        "Gene Name": gene,
                        "Protein Name": protein_name,
                        "PDB ID": self._hyperlink(
                            f"https://www.rcsb.org/structure/{pdb_id}", pdb_id
                        ),
                        "Type": issue_type,
                        "Residus": f"{start}-{end}",
                        "Length": end - start + 1,
                        "Secondary Structure": " ".join(
                            ss_map.get(residue, "C")
                            for residue in range(start, end + 1)
                        ),
                    }
                )

            details = get_pdb_details(pdb_id)
            pdb_file = out_dir / f"{gene}_{upid}_{pdb_id}.pdb"
            smalls = parse_small_molecules(str(pdb_file)) if pdb_file.exists() else []
            pmid = details.get("pubmed_id", "")
            pm_cell = (
                self._hyperlink(f"https://pubmed.ncbi.nlm.nih.gov/{pmid}", pmid)
                if pmid
                else ""
            )
            pdb_detail_rows.append(
                {
                    "UniProt ID": self._hyperlink(
                        f"https://www.uniprot.org/uniprotkb/{upid}/entry", upid
                    ),
                    "Gene Name": gene,
                    "Protein Name": protein_name,
                    "PDB ID": self._hyperlink(
                        f"https://www.rcsb.org/structure/{pdb_id}", pdb_id
                    ),
                    "PDB Source": pdb_source,
                    "Repaired": repaired,
                    "Experimental Method": details.get("method", ""),
                    "Resolution (Angstrom)": details.get("resolution", ""),
                    "Determination Methodology": details.get("methodology", ""),
                    "Conformation": details.get("conformation", ""),
                    "Environment": details.get("environment", ""),
                    "Polymer Entity Count": details.get("count", ""),
                    "Entity Descriptions": details.get("descs", ""),
                    "Chains": details.get("chains", ""),
                    "Entity Organisms": details.get("orgs", ""),
                    "PubMed ID": pm_cell,
                    "Small Molecules": ", ".join(smalls),
                }
            )

        return structural_rows, pdb_detail_rows

    def _ingest_alphafold_structures(
        self,
        upid: str,
        gene: str,
        protein_name: str,
        pred_id: Optional[str],
        partner_ids: Sequence[str],
    ) -> tuple[List[Dict[str, str]], Dict[str, int]]:
        """Collect AlphaFold-derived rows and partner length updates."""
        pdb_detail_rows: List[Dict[str, str]] = []
        prot_len_updates: Dict[str, int] = {}

        if not pred_id:
            return pdb_detail_rows, prot_len_updates

        pred_dir = self.cfg.pred_dir / f"{gene}_{upid}"
        pred_dir.mkdir(parents=True, exist_ok=True)

        af_pdb_path = download_alphafold_pdb(
            upid,
            str(pred_dir / f"{gene}_{upid}_{pred_id}.pdb"),
        )
        if af_pdb_path:
            self._set_pdb_source(pred_id or upid, Path(af_pdb_path), "AlphaFold")
        repaired = False
        if af_pdb_path and (self.cfg.use_repair or self.cfg.use_pdb_fixer):
            pdb_path = Path(af_pdb_path)
            try:
                repaired = fix_structure(
                    pdb_path,
                    pdb_path,
                    auto_install=self.cfg.auto_install_repair,
                )
                if repaired:
                    logging.info(f"PDBFixer applied to {pdb_path}")
            except Exception as exc:
                logging.error(f"PDBFixer failed for {pdb_path}: {exc}")

        partner_id_list = [p for p in partner_ids if p]
        for partner_id in partner_id_list:
            try:
                partner_json = fetch_uniprot_json(partner_id)
            except Exception as exc:
                self.log.warning(
                    "Failed to fetch partner %s for %s: %s", partner_id, upid, exc
                )
                continue

            partner_info = parse_uniprot_data(partner_json)
            partner_gene = partner_info.get("gene") or partner_id
            prot_len_updates[f"{partner_gene}_{partner_id}"] = len(
                partner_info.get("sequence", "")
            )
            partner_pred = get_alphafold_pdb_info(partner_id)
            if partner_pred:
                partner_path = pred_dir / f"{gene}_{partner_id}_{partner_pred}.pdb"
                download_alphafold_pdb(
                    partner_id,
                    str(partner_path),
                )
                self._set_pdb_source(partner_pred, partner_path, "AlphaFold")

        pred_details = get_alphafold_details(upid, protein_name)
        pdb_detail_rows.append(
            {
                "UniProt ID": self._hyperlink(
                    f"https://www.uniprot.org/uniprotkb/{upid}/entry", upid
                ),
                "Gene Name": gene,
                "Protein Name": protein_name,
                "PDB ID": self._hyperlink(
                    f"https://alphafold.ebi.ac.uk/entry/{upid}", pred_id
                ),
                "PDB Source": "AlphaFold",
                "Repaired": repaired,
                "Experimental Method": pred_details.get("method", ""),
                "Resolution (Angstrom)": pred_details.get("resolution", ""),
                "Determination Methodology": pred_details.get("methodology", ""),
                "Conformation": pred_details.get("conformation", ""),
                "Environment": pred_details.get("environment", ""),
                "Polymer Entity Count": pred_details.get("count", ""),
                "Entity Descriptions": pred_details.get("descs", ""),
                "Chains": pred_details.get("chains", ""),
                "Entity Organisms": pred_details.get("orgs", ""),
                "PubMed ID": "",
                "Small Molecules": "",
            }
        )

        return pdb_detail_rows, prot_len_updates

    # ------------------------------------------------------------------ #
    # Alignment and orientation
    # ------------------------------------------------------------------ #
    def _run_alignment(self) -> None:
        aligned_dir = self.cfg.aligned_dir
        aligned_dir.mkdir(parents=True, exist_ok=True)

        tasks: list[tuple[str, str, str, int, int, str, list[str], str]] = []
        for gene_dir in sorted(self.cfg.pred_dir.iterdir()):
            if not gene_dir.is_dir():
                continue
            try:
                gene_name, upid = gene_dir.name.split("_", 1)
            except ValueError:
                self.log.warning(
                    "Unexpected directory name in predictions: %s", gene_dir
                )
                continue

            prot_len = self._prot_len_map.get(gene_dir.name)
            if prot_len is None:
                self.log.warning(
                    "No protein length stored for %s; skipping", gene_dir.name
                )
                continue

            exp_dir = self.cfg.exp_dir / gene_dir.name
            if not exp_dir.is_dir():
                self.log.debug("No experimental directory for %s", gene_dir.name)
                continue

            partner_ids = list(self._partner_ids_map.get(gene_dir.name, []))
            for exp_file in exp_dir.glob("*.pdb"):
                tasks.append(
                    (
                        str(gene_dir),
                        str(exp_file),
                        gene_name,
                        int(prot_len),
                        int(self.cfg.window),
                        str(aligned_dir),
                        partner_ids,
                        upid,
                    )
                )

        if not tasks:
            return

        results: list[dict] = []
        max_workers = min(len(tasks), self.cfg.max_workers)
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            future_map = {
                executor.submit(_alignment_worker, task): task for task in tasks
            }
            for future in as_completed(future_map):
                task = future_map[future]
                try:
                    res = future.result()
                except Exception as exc:  # pragma: no cover - defensive logging
                    self.log.error(
                        "Alignment failed for %s vs %s: %s",
                        task[0],
                        task[1],
                        exc,
                    )
                    continue
                if res:
                    self._enrich_alignment_record(res)
                    results.append(res)

        self._alignment_summary.extend(results)

    def _enrich_alignment_record(self, record: Dict[str, object]) -> None:
        raw_id = record.get("pdb_id_raw")
        if isinstance(raw_id, str):
            protonated = self._protonation_map.get(raw_id.upper())
            if protonated is not None:
                record["protonated"] = "Yes" if protonated else "No"
            source = self._pdb_source_map.get(raw_id.upper())
            if source:
                record["pdb_source"] = source

    def _set_pdb_source(self, identifier: str, path: Path | None, source: str) -> None:
        if not source:
            return
        with self._lock:
            if identifier:
                self._pdb_source_map[identifier.upper()] = source
            if path is not None:
                self._pdb_source_map[path.stem.upper()] = source

    def _run_multi_reference_orientation(self) -> None:
        summary_df = self._data_frames.get("Alignment_Summary")
        details_df = self._data_frames.get("PDB_Details")
        if summary_df is None or summary_df.empty:
            self.log.info(
                "Skipping multi-reference orientation: no alignment summary available"
            )
            return
        if details_df is None or details_df.empty:
            self.log.info(
                "Skipping multi-reference orientation: no PDB details available"
            )
            return

        self.log.info("Running multi-reference alignment and orientation")
        
        # Rename columns to match expected format
        details_df_renamed = details_df.rename(columns={
            'UniProt ID': 'uniprot_id',
            'Gene Name': 'gene_name',
            'PDB ID': 'pdb_id',
            'Small Molecules': 'small_molecules',
            'PDB Source': 'pdb_source'
        })
        
        _, df_res = align_structures_by_groups(
            data_frames={
                "Alignment_Summary": summary_df,
                "PDB_Details": details_df_renamed,
            },
            group_by="gene_name",
            out_root=self.cfg.oriented_dir,
            cfg=self.cfg,
            orientation_fallback_fn=align_structures_by_groups.__globals__["_orient_with_fallback"],
        )

        if df_res is None:
            df_res = pd.DataFrame()

        if not df_res.empty:
            df_res["pdb_id"] = df_res["pdb_id"].astype(str)
            df_res["protonated"] = df_res["pdb_id"].map(
                lambda pid: self._protonation_map.get(str(pid).upper(), False)
            )
            orientation_map: Dict[str, tuple[Optional[str], Optional[str]]] = {
                str(row["pdb_id"]).upper(): (
                    row.get("aligned_path"),
                    row.get("orientation_method"),
                )
                for row in df_res.to_dict(orient="records")
            }
            df_res["protonated"] = df_res["protonated"].apply(
                lambda flag: "Yes" if flag else "No"
            )
        else:
            orientation_map = {}

        methods_by_uniprot: Dict[str, Set[str]] = {}
        if orientation_map:
            for record in self._alignment_summary:
                raw_id = record.get("pdb_id_raw")
                if isinstance(raw_id, str):
                    mapping = orientation_map.get(raw_id.upper())
                    if mapping:
                        aligned_path, method = mapping
                        if aligned_path:
                            record["oriented_pdb"] = aligned_path
                        record["orientation_method"] = method
                    is_protonated = bool(
                        self._protonation_map.get(raw_id.upper(), False)
                    )
                    record["protonated"] = "Yes" if is_protonated else "No"
                uni = record.get("uniprot_id")
                method = record.get("orientation_method")
                if isinstance(uni, str) and method:
                    methods_by_uniprot.setdefault(uni.upper(), set()).add(str(method))
        elif self._alignment_summary:
            for record in self._alignment_summary:
                raw_id = record.get("pdb_id_raw")
                if isinstance(raw_id, str):
                    is_protonated = bool(
                        self._protonation_map.get(raw_id.upper(), False)
                    )
                    record["protonated"] = "Yes" if is_protonated else "No"

        if self._alignment_summary:
            self._data_frames["Alignment_Summary"] = pd.DataFrame(
                self._alignment_summary
            )

        self._multi_ref_records = (
            df_res.to_dict(orient="records") if not df_res.empty else []
        )

        self._update_uniprot_orientation_methods(methods_by_uniprot)


    def _update_uniprot_orientation_methods(
        self, methods_by_uniprot: Dict[str, Set[str]]
    ) -> None:
        if not methods_by_uniprot or not self._uniprot_records:
            return
        for record in self._uniprot_records:
            label = self._extract_hyperlink_label(record.get("UniProt ID"))
            methods = methods_by_uniprot.get(label.upper()) if label else None
            if methods:
                record["Orientation Methods"] = ", ".join(sorted(methods))
        self._data_frames["UniProt Results"] = self._to_dataframe(
            self._uniprot_records, self.UNIPROT_COLUMNS
        )
        self._write_parquet(
            self._data_frames["UniProt Results"], "UniProt Results"
        )

    @staticmethod
    def _extract_hyperlink_label(value: object) -> str:
        if isinstance(value, str) and value.startswith("=HYPERLINK"):
            matches = re.findall(r'"([^"]*)"', value)
            if len(matches) >= 2:
                return matches[1]
        return str(value or "")

    # ------------------------------------------------------------------ #
    # DataFrame / storage helpers
    # ------------------------------------------------------------------ #
    def _materialise_primary_frames(self) -> None:
        self._data_frames["UniProt Results"] = self._to_dataframe(
            self._uniprot_records, self.UNIPROT_COLUMNS
        )
        self._data_frames["PDB_Details"] = self._to_dataframe(
            self._pdb_detail_records, self.PDB_DETAILS_COLUMNS
        )
        self._data_frames["Structural_Integrity"] = self._to_dataframe(
            self._structural_records, self.STRUCTURAL_COLUMNS
        )

    def _materialise_alignment_frame(self) -> None:
        summary_df = pd.DataFrame(self._alignment_summary)
        self._data_frames["Alignment_Summary"] = summary_df

    def _write_parquet_intermediates(self) -> None:
        for name in ("UniProt Results", "PDB_Details", "Structural_Integrity"):
            df = self._data_frames.get(name)
            if df is None:
                continue
            self._write_parquet(df, name)

    def _write_alignment_parquet(self) -> None:
        df = self._data_frames.get("Alignment_Summary")
        if df is not None:
            self._write_parquet(df, "Alignment_Summary")

    def _write_multi_ref_parquet(self) -> None:
        if not self._multi_ref_records:
            return
        df = pd.DataFrame(self._multi_ref_records)
        self._data_frames["Alignment_RMSD_MultiRef"] = df
        self._write_parquet(df, "Alignment_RMSD_MultiRef")

    @staticmethod
    def _stringify_for_parquet(value: object) -> str:
        if pd.api.types.is_scalar(value):
            if value is None or (
                isinstance(value, (float, np.floating)) and math.isnan(value)
            ):
                return ""
            if pd.isna(value):
                return ""
            if isinstance(value, pd.Timestamp):
                return value.isoformat()
            return str(value)
        if isinstance(value, dict):
            try:
                return json.dumps(value, ensure_ascii=False)
            except TypeError:
                return str(value)
        if isinstance(value, (list, tuple, set)):
            parts = [ProteinPipeline._stringify_for_parquet(item) for item in value]
            return ", ".join(part for part in parts if part)
        if isinstance(value, np.ndarray):
            if value.size == 0:
                return ""
            parts = [
                ProteinPipeline._stringify_for_parquet(item) for item in value.tolist()
            ]
            return ", ".join(part for part in parts if part)
        return str(value)

    @staticmethod
    def _prepare_parquet_df(df: pd.DataFrame) -> pd.DataFrame:
        prepared = df.copy()
        for column in prepared.columns:
            if prepared[column].dtype == "object":
                prepared[column] = prepared[column].apply(
                    ProteinPipeline._stringify_for_parquet
                )
        return prepared

    def _write_parquet(self, df: pd.DataFrame, name: str) -> None:
        target = self.cfg.parquet_dir / f"{name.replace(' ', '_').lower()}.parquet"
        target.parent.mkdir(parents=True, exist_ok=True)
        self._prepare_parquet_df(df).to_parquet(target, index=False)
        self.log.debug("Wrote %s (%d rows) to %s", name, len(df), target)

    def _export_excel(self) -> None:
        if not self._data_frames:
            self.log.warning("No data frames available; skipping Excel export")
            return

        ordered_names = [
            "UniProt Results",
            "PDB_Details",
            "Structural_Integrity",
            "Alignment_Summary",
            "Alignment_RMSD_MultiRef",
        ]

        entries = [
            (name, self._data_frames.get(name))
            for name in ordered_names
            if self._data_frames.get(name) is not None
            and not self._data_frames.get(name).empty
        ]

        if not entries:
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws.append(["message"])
            ws.append(["No data generated"])
            wb.save(self.cfg.output_excel)
            self.log.info("Excel export skipped: no data generated")
            return

        engine = "xlsxwriter" if self.cfg.use_xlsxwriter else "openpyxl"
        if self.cfg.use_xlsxwriter:
            try:
                import xlsxwriter  # noqa: F401
                logging.info("Using XlsxWriter for initial Excel export.")
            except ImportError:
                self.log.warning(
                    "xlsxwriter not installed, falling back to openpyxl for Excel export."
                )
                engine = "openpyxl"
        try:
            with pd.ExcelWriter(self.cfg.output_excel, engine=engine) as writer:
                for name, df in entries:
                    df.to_excel(writer, sheet_name=name, index=False)
        except (PermissionError, IOError) as exc:
            self.log.error(
                "Unable to write %s (is the file open elsewhere?): %s",
                self.cfg.output_excel,
                exc,
            )
            raise

        wb = load_workbook(self.cfg.output_excel)
        for sheet_name, columns in (
            ("UniProt Results", (1, 2)),
            ("PDB_Details", (1, 4, 16)),
            ("Structural_Integrity", (1, 4)),
            ("Alignment_Summary", (1,)),
        ):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for col_idx in columns:
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    style_link(row[0], self.link_font)

            if sheet_name == "Alignment_Summary":
                status_col = None
                for i, cell in enumerate(ws[1], start=1):
                    if cell.value == "status":
                        status_col = i
                        break
                if status_col:
                    for row in ws.iter_rows(
                        min_row=2, min_col=status_col, max_col=status_col
                    ):
                        cell = row[0]
                        value = str(cell.value).lower()
                        if value == "ok":
                            cell.fill = PatternFill("solid", fgColor="C6EFCE")
                        else:
                            cell.fill = PatternFill("solid", fgColor="FFC7CE")

        wb.save(self.cfg.output_excel)

    # ------------------------------------------------------------------ #
    # Utilities
    # ------------------------------------------------------------------ #
    @staticmethod
    def _hyperlink(url: str | None, label: object) -> str:
        safe_url = str(url or "").replace('"', '""')
        safe_label = str(label or "").replace('"', '""')
        return f'=HYPERLINK("{safe_url}", "{safe_label}")'

    @staticmethod
    def _to_dataframe(
        records: List[Dict[str, object]], columns: List[str]
    ) -> pd.DataFrame:
        if not records:
            return pd.DataFrame(columns=columns)
        return pd.DataFrame(records, columns=columns)


# --------------------------------------------------------------------------- #
# CLI entry point
# --------------------------------------------------------------------------- #


def build_arg_parser() -> argparse.ArgumentParser:
    """Create the command-line parser for the pipeline CLI."""
    parser = argparse.ArgumentParser("Protein structure pipeline")
    parser.add_argument("--base", default=None, help="Base working directory")
    parser.add_argument("--input", default=None, help="Path to UniProt input template")
    parser.add_argument("--output", default=None, help="Output Excel filename")
    parser.add_argument("--window", type=int, default=5, help="Sliding window size")
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbose logging")
    parser.add_argument("--ids", nargs="+", help="Process explicit UniProt IDs")
    return parser


def main() -> None:
    """Entry point for running the pipeline from the command line."""
    parser = build_arg_parser()
    args = parser.parse_args()
    config = PipelineConfig.from_args(
        base=args.base,
        input_template=args.input,
        output_excel=args.output,
        window=args.window,
    )
    pipeline = ProteinPipeline(config=config, ids=args.ids, verbose=args.verbose)
    pipeline.run()


if __name__ == "__main__":
    main()
